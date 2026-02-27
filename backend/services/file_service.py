# NEXUS AI - File Service
"""
File operations service with PDF generation (reportlab), Excel generation
(openpyxl), Markdown report writing, file watching (watchdog), and backup
management for the NEXUS AI OS.
"""

import asyncio
import hashlib
import json
import os
import shutil
import time
import uuid
import zipfile
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional, Tuple, Union

from loguru import logger

from core.config import NexusSettings, settings
from core.events import Event, EventBus, EventCategory, EventPriority, event_bus
from core.logger import nexus_logger


class FileInfo:
    """Metadata about a filesystem object."""

    def __init__(self, path: str):
        self.path: str = path
        p = Path(path)
        self.name: str = p.name
        self.extension: str = p.suffix
        self.exists: bool = p.exists()
        self.is_file: bool = p.is_file() if p.exists() else False
        self.is_dir: bool = p.is_dir() if p.exists() else False
        self.size_bytes: int = p.stat().st_size if p.is_file() else 0
        self.created: Optional[datetime] = (
            datetime.fromtimestamp(p.stat().st_ctime) if p.exists() else None
        )
        self.modified: Optional[datetime] = (
            datetime.fromtimestamp(p.stat().st_mtime) if p.exists() else None
        )

    def to_dict(self) -> Dict[str, Any]:
        """Serialize file info to dictionary."""
        return {
            "path": self.path,
            "name": self.name,
            "extension": self.extension,
            "exists": self.exists,
            "is_file": self.is_file,
            "is_dir": self.is_dir,
            "size_bytes": self.size_bytes,
            "size_human": self._human_size(self.size_bytes),
            "created": self.created.isoformat() if self.created else None,
            "modified": self.modified.isoformat() if self.modified else None,
        }

    @staticmethod
    def _human_size(size: int) -> str:
        """Convert byte size to human readable string."""
        for unit in ("B", "KB", "MB", "GB", "TB"):
            if abs(size) < 1024.0:
                return f"{size:.1f} {unit}"
            size /= 1024.0  # type: ignore
        return f"{size:.1f} PB"


class FileWatcher:
    """File system watcher using watchdog for real-time change detection."""

    def __init__(self, watch_paths: Optional[List[str]] = None):
        self._watch_paths: List[str] = watch_paths or []
        self._observer: Any = None
        self._running: bool = False
        self._callbacks: List[Callable] = []
        self._event_count: int = 0

    def add_callback(self, callback: Callable) -> None:
        """Register a callback for file change events."""
        self._callbacks.append(callback)

    async def start(self) -> None:
        """Start watching the configured paths."""
        if self._running:
            return
        try:
            from watchdog.observers import Observer
            from watchdog.events import FileSystemEventHandler, FileSystemEvent

            service_ref = self

            class NexusEventHandler(FileSystemEventHandler):
                def on_any_event(self, event: FileSystemEvent) -> None:
                    service_ref._event_count += 1
                    for cb in service_ref._callbacks:
                        try:
                            cb({
                                "type": event.event_type,
                                "path": event.src_path,
                                "is_directory": event.is_directory,
                                "timestamp": datetime.utcnow().isoformat(),
                            })
                        except Exception as exc:
                            logger.error(f"File watcher callback error: {exc}")

            self._observer = Observer()
            handler = NexusEventHandler()
            for path in self._watch_paths:
                if os.path.exists(path):
                    self._observer.schedule(handler, path, recursive=True)
                    logger.info(f"Watching directory: {path}")

            self._observer.start()
            self._running = True
            logger.info("File watcher started")
        except Exception as exc:
            logger.error(f"File watcher start error: {exc}")

    async def stop(self) -> None:
        """Stop the file watcher."""
        if self._observer and self._running:
            self._observer.stop()
            self._observer.join(timeout=5)
            self._running = False
            logger.info("File watcher stopped")

    @property
    def event_count(self) -> int:
        """Total number of file system events observed."""
        return self._event_count


class FileService:
    """
    Comprehensive file operations service for NEXUS AI.

    Provides:
    - PDF generation using reportlab
    - Excel spreadsheet generation using openpyxl
    - Markdown report writing
    - File watching via watchdog
    - Backup creation and management
    - File search and metadata extraction
    - Directory management
    - File hashing and integrity checking
    """

    def __init__(self, config: Optional[NexusSettings] = None,
                 event_bus_instance: Optional[EventBus] = None):
        self._config: NexusSettings = config or settings
        self._event_bus: EventBus = event_bus_instance or event_bus
        self._data_dir: Path = Path(self._config.database.db_path).parent
        self._reports_dir: Path = self._data_dir / "reports"
        self._backup_dir: Path = Path(self._config.database.backup_path)
        self._exports_dir: Path = self._data_dir / "exports"
        self._watcher: FileWatcher = FileWatcher()
        self._initialized: bool = False
        self._files_created: int = 0
        self._files_deleted: int = 0
        self._backups_created: int = 0

    # ------------------------------------------------------------------
    # Lifecycle
    # ------------------------------------------------------------------

    async def initialize(self) -> None:
        """Initialize directories and file watcher."""
        try:
            logger.info("Initializing FileService...")
            for d in [self._reports_dir, self._backup_dir, self._exports_dir]:
                d.mkdir(parents=True, exist_ok=True)

            self._watcher = FileWatcher(
                watch_paths=[str(self._data_dir)],
            )
            self._watcher.add_callback(self._on_file_change)
            await self._watcher.start()

            self._initialized = True
            await self._event_bus.emit(
                "file.initialized",
                {"reports_dir": str(self._reports_dir), "backup_dir": str(self._backup_dir)},
                source="file_service",
                category=EventCategory.SYSTEM,
            )
            logger.info("FileService initialized")
        except Exception as exc:
            logger.error(f"FileService initialization failed: {exc}")
            self._initialized = True

    async def shutdown(self) -> None:
        """Stop file watcher and clean up."""
        try:
            logger.info("Shutting down FileService...")
            await self._watcher.stop()
            self._initialized = False
            logger.info("FileService shut down complete")
        except Exception as exc:
            logger.error(f"Error during FileService shutdown: {exc}")

    def _on_file_change(self, event_data: Dict[str, Any]) -> None:
        """Handle file system change events from the watcher."""
        event_type = event_data.get("type", "unknown")
        path = event_data.get("path", "")
        logger.debug(f"File change: {event_type} — {path}")

    # ------------------------------------------------------------------
    # PDF Generation
    # ------------------------------------------------------------------

    async def generate_pdf(
        self,
        title: str,
        content: List[Dict[str, Any]],
        output_path: Optional[str] = None,
        metadata: Optional[Dict[str, str]] = None,
    ) -> str:
        """
        Generate a PDF report using reportlab.

        Args:
            title: Report title.
            content: List of content blocks, each a dict with 'type' and 'data'.
                     Supported types: 'heading', 'paragraph', 'table', 'spacer'.
            output_path: Optional output file path. Auto-generated if not provided.
            metadata: Optional PDF metadata (author, subject, etc.).

        Returns:
            Path to the generated PDF.
        """
        if not output_path:
            filename = f"report_{uuid.uuid4().hex[:8]}_{int(time.time())}.pdf"
            output_path = str(self._reports_dir / filename)

        def _build_pdf() -> str:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import letter
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            from reportlab.platypus import (
                SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
                PageBreak,
            )

            Path(output_path).parent.mkdir(parents=True, exist_ok=True)
            doc = SimpleDocTemplate(
                output_path,
                pagesize=letter,
                rightMargin=72,
                leftMargin=72,
                topMargin=72,
                bottomMargin=72,
            )

            if metadata:
                doc.title = metadata.get("title", title)
                doc.author = metadata.get("author", "NEXUS AI")
                doc.subject = metadata.get("subject", "")

            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                "NexusTitle",
                parent=styles["Title"],
                fontSize=24,
                spaceAfter=30,
                textColor=colors.HexColor("#1a237e"),
            )
            heading_style = ParagraphStyle(
                "NexusHeading",
                parent=styles["Heading1"],
                fontSize=16,
                spaceAfter=12,
                textColor=colors.HexColor("#283593"),
            )
            body_style = styles["Normal"]

            elements = []
            elements.append(Paragraph(title, title_style))
            elements.append(Paragraph(
                f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC')}",
                body_style,
            ))
            elements.append(Spacer(1, 0.3 * inch))

            for block in content:
                block_type = block.get("type", "paragraph")
                data = block.get("data", "")

                if block_type == "heading":
                    elements.append(Paragraph(str(data), heading_style))
                elif block_type == "paragraph":
                    elements.append(Paragraph(str(data), body_style))
                    elements.append(Spacer(1, 0.15 * inch))
                elif block_type == "table":
                    table_data = data if isinstance(data, list) else []
                    if table_data:
                        t = Table(table_data)
                        t.setStyle(TableStyle([
                            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1a237e")),
                            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                            ("FONTSIZE", (0, 0), (-1, 0), 11),
                            ("BOTTOMPADDING", (0, 0), (-1, 0), 10),
                            ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#e8eaf6")),
                            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                            ("ROWBACKGROUNDS", (0, 1), (-1, -1),
                             [colors.white, colors.HexColor("#e8eaf6")]),
                        ]))
                        elements.append(t)
                        elements.append(Spacer(1, 0.2 * inch))
                elif block_type == "spacer":
                    height = float(data) if data else 0.3
                    elements.append(Spacer(1, height * inch))
                elif block_type == "page_break":
                    elements.append(PageBreak())

            doc.build(elements)
            return output_path

        loop = asyncio.get_running_loop()
        result_path = await loop.run_in_executor(None, _build_pdf)
        self._files_created += 1

        nexus_logger.log_activity(
            "pdf_generated", f"PDF report: {title}",
            metadata={"path": result_path},
        )
        await self._event_bus.emit(
            "file.pdf_generated",
            {"title": title, "path": result_path},
            source="file_service",
            category=EventCategory.SYSTEM,
        )
        logger.info(f"PDF generated: {result_path}")
        return result_path

    # ------------------------------------------------------------------
    # Excel Generation
    # ------------------------------------------------------------------

    async def generate_excel(
        self,
        title: str,
        sheets: Dict[str, List[List[Any]]],
        output_path: Optional[str] = None,
        column_widths: Optional[Dict[str, List[int]]] = None,
    ) -> str:
        """
        Generate an Excel workbook using openpyxl.

        Args:
            title: Workbook title (used for filename if output_path not given).
            sheets: Dict mapping sheet names to row data.
                    First row of each sheet is treated as headers.
            output_path: Optional output path.
            column_widths: Optional dict mapping sheet names to list of column widths.

        Returns:
            Path to the generated Excel file.
        """
        if not output_path:
            safe_title = "".join(c for c in title if c.isalnum() or c in " _-")[:50]
            filename = f"{safe_title}_{int(time.time())}.xlsx"
            output_path = str(self._exports_dir / filename)

        def _build_excel() -> str:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            Path(output_path).parent.mkdir(parents=True, exist_ok=True)
            wb = Workbook()
            wb.remove(wb.active)

            header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="1A237E", end_color="1A237E", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            cell_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )
            alt_fill = PatternFill(start_color="E8EAF6", end_color="E8EAF6", fill_type="solid")

            for sheet_name, rows in sheets.items():
                ws = wb.create_sheet(title=sheet_name[:31])
                for row_idx, row_data in enumerate(rows, 1):
                    for col_idx, value in enumerate(row_data, 1):
                        cell = ws.cell(row=row_idx, column=col_idx, value=value)
                        cell.border = cell_border

                        if row_idx == 1:
                            cell.font = header_font
                            cell.fill = header_fill
                            cell.alignment = header_alignment
                        else:
                            if row_idx % 2 == 0:
                                cell.fill = alt_fill
                            cell.alignment = Alignment(horizontal="left")

                widths = (column_widths or {}).get(sheet_name)
                if widths:
                    for col_idx, width in enumerate(widths, 1):
                        ws.column_dimensions[get_column_letter(col_idx)].width = width
                else:
                    for col_idx in range(1, ws.max_column + 1):
                        max_len = 0
                        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
                            for cell in row:
                                if cell.value:
                                    max_len = max(max_len, len(str(cell.value)))
                        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 50)

                ws.auto_filter.ref = ws.dimensions

            wb.save(output_path)
            return output_path

        loop = asyncio.get_running_loop()
        result_path = await loop.run_in_executor(None, _build_excel)
        self._files_created += 1

        nexus_logger.log_activity(
            "excel_generated", f"Excel workbook: {title}",
            metadata={"path": result_path, "sheets": list(sheets.keys())},
        )
        await self._event_bus.emit(
            "file.excel_generated",
            {"title": title, "path": result_path, "sheets": list(sheets.keys())},
            source="file_service",
            category=EventCategory.SYSTEM,
        )
        logger.info(f"Excel generated: {result_path}")
        return result_path

    # ------------------------------------------------------------------
    # Markdown Reports
    # ------------------------------------------------------------------

    async def generate_markdown(
        self,
        title: str,
        sections: List[Dict[str, Any]],
        output_path: Optional[str] = None,
    ) -> str:
        """
        Generate a Markdown report.

        Args:
            title: Report title.
            sections: List of section dicts with 'heading', 'content', and optional 'level'.
            output_path: Optional output path.

        Returns:
            Path to the generated Markdown file.
        """
        if not output_path:
            safe_title = "".join(c for c in title if c.isalnum() or c in " _-")[:50]
            filename = f"{safe_title}_{int(time.time())}.md"
            output_path = str(self._reports_dir / filename)

        def _build_md() -> str:
            Path(output_path).parent.mkdir(parents=True, exist_ok=True)
            lines: List[str] = []
            lines.append(f"# {title}")
            lines.append("")
            lines.append(f"*Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC')}*")
            lines.append(f"*By: NEXUS AI*")
            lines.append("")
            lines.append("---")
            lines.append("")

            for section in sections:
                level = section.get("level", 2)
                heading = section.get("heading", "")
                content = section.get("content", "")
                prefix = "#" * level

                if heading:
                    lines.append(f"{prefix} {heading}")
                    lines.append("")

                if isinstance(content, str):
                    lines.append(content)
                    lines.append("")
                elif isinstance(content, list):
                    if content and isinstance(content[0], list):
                        # Table
                        if len(content) > 0:
                            headers = content[0]
                            lines.append("| " + " | ".join(str(h) for h in headers) + " |")
                            lines.append("| " + " | ".join("---" for _ in headers) + " |")
                            for row in content[1:]:
                                lines.append("| " + " | ".join(str(c) for c in row) + " |")
                            lines.append("")
                    else:
                        # Bullet list
                        for item in content:
                            lines.append(f"- {item}")
                        lines.append("")
                elif isinstance(content, dict):
                    for key, value in content.items():
                        lines.append(f"- **{key}**: {value}")
                    lines.append("")

            md_text = "\n".join(lines)
            with open(output_path, "w", encoding="utf-8") as f:
                f.write(md_text)
            return output_path

        loop = asyncio.get_running_loop()
        result_path = await loop.run_in_executor(None, _build_md)
        self._files_created += 1

        await self._event_bus.emit(
            "file.markdown_generated",
            {"title": title, "path": result_path},
            source="file_service",
            category=EventCategory.SYSTEM,
        )
        logger.info(f"Markdown report generated: {result_path}")
        return result_path

    # ------------------------------------------------------------------
    # File Operations
    # ------------------------------------------------------------------

    async def read_file(self, file_path: str, encoding: str = "utf-8") -> str:
        """
        Read a text file asynchronously.

        Args:
            file_path: Path to the file.
            encoding: File encoding.

        Returns:
            File contents as string.
        """
        def _read() -> str:
            with open(file_path, "r", encoding=encoding) as f:
                return f.read()

        loop = asyncio.get_running_loop()
        return await loop.run_in_executor(None, _read)

    async def write_file(self, file_path: str, content: str,
                         encoding: str = "utf-8") -> str:
        """
        Write content to a text file.

        Args:
            file_path: Target file path.
            content: Text content to write.
            encoding: File encoding.

        Returns:
            The file path.
        """
        def _write() -> str:
            Path(file_path).parent.mkdir(parents=True, exist_ok=True)
            with open(file_path, "w", encoding=encoding) as f:
                f.write(content)
            return file_path

        loop = asyncio.get_running_loop()
        result = await loop.run_in_executor(None, _write)
        self._files_created += 1
        return result

    async def read_binary(self, file_path: str) -> bytes:
        """Read a binary file."""
        def _read() -> bytes:
            with open(file_path, "rb") as f:
                return f.read()

        loop = asyncio.get_running_loop()
        return await loop.run_in_executor(None, _read)

    async def write_binary(self, file_path: str, data: bytes) -> str:
        """Write binary data to a file."""
        def _write() -> str:
            Path(file_path).parent.mkdir(parents=True, exist_ok=True)
            with open(file_path, "wb") as f:
                f.write(data)
            return file_path

        loop = asyncio.get_running_loop()
        result = await loop.run_in_executor(None, _write)
        self._files_created += 1
        return result

    async def delete_file(self, file_path: str) -> bool:
        """
        Delete a file.

        Args:
            file_path: Path to file to delete.

        Returns:
            True if deletion was successful.
        """
        def _delete() -> bool:
            try:
                os.remove(file_path)
                return True
            except OSError as exc:
                logger.error(f"File deletion error: {exc}")
                return False

        loop = asyncio.get_running_loop()
        result = await loop.run_in_executor(None, _delete)
        if result:
            self._files_deleted += 1
        return result

    async def copy_file(self, src: str, dst: str) -> str:
        """Copy a file from src to dst."""
        def _copy() -> str:
            Path(dst).parent.mkdir(parents=True, exist_ok=True)
            shutil.copy2(src, dst)
            return dst

        loop = asyncio.get_running_loop()
        return await loop.run_in_executor(None, _copy)

    async def move_file(self, src: str, dst: str) -> str:
        """Move a file from src to dst."""
        def _move() -> str:
            Path(dst).parent.mkdir(parents=True, exist_ok=True)
            shutil.move(src, dst)
            return dst

        loop = asyncio.get_running_loop()
        return await loop.run_in_executor(None, _move)

    async def get_file_info(self, file_path: str) -> Dict[str, Any]:
        """Get metadata about a file."""
        def _info() -> Dict[str, Any]:
            return FileInfo(file_path).to_dict()

        loop = asyncio.get_running_loop()
        return await loop.run_in_executor(None, _info)

    async def list_directory(self, dir_path: str,
                             pattern: str = "*",
                             recursive: bool = False) -> List[Dict[str, Any]]:
        """
        List contents of a directory.

        Args:
            dir_path: Directory path.
            pattern: Glob pattern for filtering.
            recursive: Whether to search recursively.

        Returns:
            List of file info dicts.
        """
        def _list() -> List[Dict[str, Any]]:
            p = Path(dir_path)
            if not p.exists():
                return []
            if recursive:
                matches = list(p.rglob(pattern))
            else:
                matches = list(p.glob(pattern))
            return [FileInfo(str(m)).to_dict() for m in matches[:1000]]

        loop = asyncio.get_running_loop()
        return await loop.run_in_executor(None, _list)

    async def search_files(self, root_dir: str, keyword: str,
                           extensions: Optional[List[str]] = None) -> List[Dict[str, Any]]:
        """
        Search for files containing a keyword in their name.

        Args:
            root_dir: Root directory to search.
            keyword: Search keyword.
            extensions: Optional list of file extensions to filter.

        Returns:
            List of matching file info dicts.
        """
        def _search() -> List[Dict[str, Any]]:
            results: List[Dict[str, Any]] = []
            root = Path(root_dir)
            if not root.exists():
                return results
            for item in root.rglob("*"):
                if not item.is_file():
                    continue
                if keyword.lower() in item.name.lower():
                    if extensions:
                        if item.suffix.lower() not in [e.lower() for e in extensions]:
                            continue
                    results.append(FileInfo(str(item)).to_dict())
                    if len(results) >= 500:
                        break
            return results

        loop = asyncio.get_running_loop()
        return await loop.run_in_executor(None, _search)

    async def compute_hash(self, file_path: str, algorithm: str = "sha256") -> str:
        """
        Compute the hash of a file.

        Args:
            file_path: Path to the file.
            algorithm: Hash algorithm (md5, sha1, sha256).

        Returns:
            Hex digest string.
        """
        def _hash() -> str:
            h = hashlib.new(algorithm)
            with open(file_path, "rb") as f:
                while True:
                    chunk = f.read(8192)
                    if not chunk:
                        break
                    h.update(chunk)
            return h.hexdigest()

        loop = asyncio.get_running_loop()
        return await loop.run_in_executor(None, _hash)

    # ------------------------------------------------------------------
    # Backup Management
    # ------------------------------------------------------------------

    async def create_backup(
        self,
        source_paths: List[str],
        backup_name: Optional[str] = None,
        compress: bool = True,
    ) -> str:
        """
        Create a backup archive of specified paths.

        Args:
            source_paths: List of file/directory paths to back up.
            backup_name: Optional name for the backup file.
            compress: Whether to use ZIP compression.

        Returns:
            Path to the backup file.
        """
        if not backup_name:
            timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
            backup_name = f"nexus_backup_{timestamp}"

        backup_path = str(self._backup_dir / f"{backup_name}.zip")

        def _create() -> str:
            Path(backup_path).parent.mkdir(parents=True, exist_ok=True)
            compression = zipfile.ZIP_DEFLATED if compress else zipfile.ZIP_STORED
            with zipfile.ZipFile(backup_path, "w", compression) as zf:
                for src in source_paths:
                    src_path = Path(src)
                    if src_path.is_file():
                        zf.write(src, src_path.name)
                    elif src_path.is_dir():
                        for item in src_path.rglob("*"):
                            if item.is_file():
                                arcname = str(item.relative_to(src_path.parent))
                                zf.write(str(item), arcname)
            return backup_path

        loop = asyncio.get_running_loop()
        result = await loop.run_in_executor(None, _create)
        self._backups_created += 1

        file_info = await self.get_file_info(result)
        nexus_logger.log_activity(
            "backup_created", f"Backup: {backup_name}",
            metadata={"path": result, "size": file_info.get("size_human", "")},
        )
        await self._event_bus.emit(
            "file.backup_created",
            {"path": result, "name": backup_name, "sources": source_paths},
            source="file_service",
            category=EventCategory.SYSTEM,
            priority=EventPriority.HIGH,
        )
        logger.info(f"Backup created: {result}")
        return result

    async def restore_backup(self, backup_path: str, restore_dir: str) -> bool:
        """
        Restore files from a backup archive.

        Args:
            backup_path: Path to the backup ZIP.
            restore_dir: Directory to extract into.

        Returns:
            True if restoration succeeded.
        """
        def _restore() -> bool:
            try:
                Path(restore_dir).mkdir(parents=True, exist_ok=True)
                with zipfile.ZipFile(backup_path, "r") as zf:
                    zf.extractall(restore_dir)
                return True
            except Exception as exc:
                logger.error(f"Backup restore error: {exc}")
                return False

        loop = asyncio.get_running_loop()
        result = await loop.run_in_executor(None, _restore)
        if result:
            logger.info(f"Backup restored: {backup_path} -> {restore_dir}")
        return result

    async def list_backups(self) -> List[Dict[str, Any]]:
        """List all available backup files."""
        return await self.list_directory(str(self._backup_dir), "*.zip")

    async def cleanup_old_backups(self, max_age_days: int = 30,
                                  keep_minimum: int = 5) -> int:
        """
        Remove backups older than max_age_days, keeping at least keep_minimum.

        Args:
            max_age_days: Maximum age in days.
            keep_minimum: Minimum number of backups to keep.

        Returns:
            Number of backups deleted.
        """
        def _cleanup() -> int:
            backup_dir = Path(self._backup_dir)
            backups = sorted(backup_dir.glob("*.zip"), key=lambda p: p.stat().st_mtime)
            if len(backups) <= keep_minimum:
                return 0

            cutoff = datetime.utcnow().timestamp() - (max_age_days * 86400)
            deleted = 0
            for backup in backups:
                if len(backups) - deleted <= keep_minimum:
                    break
                if backup.stat().st_mtime < cutoff:
                    try:
                        backup.unlink()
                        deleted += 1
                    except OSError as exc:
                        logger.error(f"Failed to delete old backup {backup}: {exc}")
            return deleted

        loop = asyncio.get_running_loop()
        deleted = await loop.run_in_executor(None, _cleanup)
        if deleted > 0:
            logger.info(f"Cleaned up {deleted} old backups")
        return deleted

    # ------------------------------------------------------------------
    # Directory Management
    # ------------------------------------------------------------------

    async def ensure_directory(self, dir_path: str) -> str:
        """Ensure a directory exists, creating it if necessary."""
        def _ensure() -> str:
            Path(dir_path).mkdir(parents=True, exist_ok=True)
            return dir_path

        loop = asyncio.get_running_loop()
        return await loop.run_in_executor(None, _ensure)

    async def get_directory_size(self, dir_path: str) -> Dict[str, Any]:
        """
        Calculate total size of a directory.

        Returns:
            Dict with 'total_bytes', 'total_human', 'file_count'.
        """
        def _size() -> Dict[str, Any]:
            total = 0
            file_count = 0
            for item in Path(dir_path).rglob("*"):
                if item.is_file():
                    total += item.stat().st_size
                    file_count += 1
            return {
                "total_bytes": total,
                "total_human": FileInfo._human_size(total),
                "file_count": file_count,
                "path": dir_path,
            }

        loop = asyncio.get_running_loop()
        return await loop.run_in_executor(None, _size)

    # ------------------------------------------------------------------
    # Health & Stats
    # ------------------------------------------------------------------

    async def health_check(self) -> Dict[str, Any]:
        """Return file service health status."""
        return {
            "service": "file_service",
            "initialized": self._initialized,
            "reports_dir": str(self._reports_dir),
            "backup_dir": str(self._backup_dir),
            "exports_dir": str(self._exports_dir),
            "watcher_running": self._watcher._running,
            "watcher_events": self._watcher.event_count,
            "files_created": self._files_created,
            "files_deleted": self._files_deleted,
            "backups_created": self._backups_created,
        }

    def get_stats(self) -> Dict[str, Any]:
        """Return runtime statistics."""
        return {
            "initialized": self._initialized,
            "files_created": self._files_created,
            "files_deleted": self._files_deleted,
            "backups_created": self._backups_created,
            "watcher_events": self._watcher.event_count,
        }
